"""
Extract BLS OEWS wage percentile data from the all_data Excel file
and produce a compact JSON for bundling in the Chrome extension.

Usage: python scripts/build-wage-data.py /tmp/oews/oesm24all/all_data_M_2024.xlsx

Output: src/data/wages.json
"""

import json
import sys
import os
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 2:
        print("Usage: python build-wage-data.py <path_to_xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    print(f"Loading {xlsx_path}...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Find column indices from header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}

    required = ['AREA', 'AREA_TITLE', 'OCC_CODE', 'OCC_TITLE', 'O_GROUP',
                'A_PCT10', 'A_PCT25', 'A_MEDIAN', 'A_PCT75', 'A_PCT90']
    for r in required:
        if r not in col:
            print(f"Missing column: {r}")
            print(f"Available: {headers}")
            sys.exit(1)

    print("Extracting wage data...")
    data = {}  # key: "AREA|OCC_CODE" -> values
    row_count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        if row_count % 100000 == 0:
            print(f"  ...{row_count:,} rows processed")

        o_group = row[col['O_GROUP']]
        # Only include detailed occupations (not summaries)
        if o_group != 'detailed':
            skipped += 1
            continue

        area = str(row[col['AREA']]).strip()
        occ_code = str(row[col['OCC_CODE']]).strip().replace('-', '')
        occ_title = str(row[col['OCC_TITLE']]).strip()
        area_title = str(row[col['AREA_TITLE']]).strip()

        def parse_wage(val):
            if val is None:
                return None
            s = str(val).strip().replace(',', '')
            if s in ('*', '#', '**', '-', ''):
                return None
            try:
                n = float(s)
                return int(round(n)) if n > 0 else None
            except ValueError:
                return None

        p10 = parse_wage(row[col['A_PCT10']])
        p25 = parse_wage(row[col['A_PCT25']])
        p50 = parse_wage(row[col['A_MEDIAN']])
        p75 = parse_wage(row[col['A_PCT75']])
        p90 = parse_wage(row[col['A_PCT90']])

        # Must have at least median
        if p50 is None:
            skipped += 1
            continue

        key = f"{area}|{occ_code}"
        data[key] = {
            'a': area,
            'at': area_title,
            'soc': occ_code,
            'occ': occ_title,
            'p10': p10,
            'p25': p25,
            'p50': p50,
            'p75': p75,
            'p90': p90,
        }

    wb.close()

    print(f"Total rows: {row_count:,}")
    print(f"Skipped: {skipped:,}")
    print(f"Valid entries: {len(data):,}")

    # Convert to compact array format to minimize size
    # Format: [[area, soc, p10, p25, p50, p75, p90], ...]
    entries = []
    # Also build area title and occupation title lookup tables
    area_titles = {}
    occ_titles = {}

    for v in data.values():
        area = v['a']
        soc = v['soc']

        if area not in area_titles:
            area_titles[area] = v['at']
        if soc not in occ_titles:
            occ_titles[soc] = v['occ']

        entries.append([area, soc, v['p10'], v['p25'], v['p50'], v['p75'], v['p90']])

    output = {
        'version': '2024.05',
        'areas': area_titles,
        'occupations': occ_titles,
        'wages': entries,
    }

    # Write output
    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'src', 'data')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'wages.json')

    with open(out_path, 'w') as f:
        json.dump(output, f, separators=(',', ':'))

    size_mb = os.path.getsize(out_path) / (1024 * 1024)
    print(f"\nOutput: {out_path}")
    print(f"Size: {size_mb:.1f} MB")
    print(f"Areas: {len(area_titles)}")
    print(f"Occupations: {len(occ_titles)}")
    print(f"Wage entries: {len(entries):,}")

if __name__ == '__main__':
    main()
